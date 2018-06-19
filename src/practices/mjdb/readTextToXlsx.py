#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time    : 2018/6/14 15:39
# @Author  : liugang9
# @Email   : mlcc330@hotmail.com
# @File    : readTextToXlsx.py
# @Software: PyCharm
# @license: Apache Licence
# @contact: 3323202070@qq.com

"""

"""

import openpyxl,xlwt,os
from openpyxl import load_workbook
from openpyxl import Workbook

# def createwb(wbname):
#     wb = openpyxl.Workbook()
#     wb.save(filename=wbname)
#     print('new a .xlsx is successful')
# sheet = wb.add_sheet('data')

# def saveToExcel(data,fileds,sheetname,wbname):
#     wb = openpyxl.load_workbook(filename=wbname)
#     sheet = wb.active
#     sheet.title = sheetname
#     field = 1
#     for field in range(1,len(fileds)+1):
#         _= sheet.cell

# input_filename = os.path.abspath('a.txt')
# output_filename = os.path.abspath('out.xls')
input_filename = "E:\\Developer\\JetBrains\\PyCharm 2017.3.3\\workspaces\\mp\\test\\SourceData\\a.txt"
output_filename = "E:\\Developer\\JetBrains\\PyCharm 2017.3.3\\workspaces\\mp\\test\\TargetData\\out.xls"

def readTxtFile(input_filename):
    with open(input_filename, 'r', encoding='utf-8') as files:
        lines = files.readlines()
    return lines

def writeXls2003File(lines):
    wb = xlwt.Workbook(encoding='utf-8', style_compression=0)
    sheet = wb.add_sheet('report', cell_overwrite_ok=True)
    if type(lines).__name__ == 'list':
        for i in range(len(lines)):
            sheet.write(i,0,lines[i])
    else:
        print('the input parameter is wrong!.')
    wb.save(os.path.abspath(output_filename))

def main():
    writeXls2003File(readTxtFile(input_filename))
    pass

if __name__ == '__main__':
    main()
    pass
